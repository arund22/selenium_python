import openpyxl
import xlrd
import xlutils
from selenium import webdriver
from openpyxl import load_workbook
import os


def data_set_up(testcasename):
    path = '/Users/wipro/Documents/Selenium_Project/E-commerce/utilities/datadriven/Test_Data.xlsx'
    workbook = xlrd.open_workbook(path)
    sheet = workbook.sheet_by_index(0)
    #print(sheet.nrows)
    #print(sheet.ncols)
    nrows = sheet.nrows
    ncols = sheet.ncols
    data = []

    for i in range(1,nrows):
        #print(sheet.cell_value(i,0))
        if sheet.cell_value(i,0) == testcasename:
            for j in range(1,ncols):
                data.append(sheet.cell_value(i,j))
        print(" ")

    return data

#print(sheet.cell_value(1,0))


#workbook = openpyxl.load_workbook(path)
#sheetname = 'Test_Name'
#sheet = openpyxl.load_workbook(sheetname)
#rowno = sheet.max_row
#colno = sheet.max_column
#print(rowno)
#print(colno)
